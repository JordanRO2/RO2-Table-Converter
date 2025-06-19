"""Excel (.xlsx) File Processor for Ragnarok Online 2 CT Format"""

import logging
from pathlib import Path
from typing import List, Any, Dict, Type, Union, TypedDict
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter

# Configure logging
logger = logging.getLogger(__name__)


class CTTypeFormat(TypedDict):
    """Type definition for CT format specifications."""
    num_format: str
    python_type: Type[Union[int, float, bool, str]]


class ValidationResult(TypedDict):
    """Type definition for validation results."""
    valid: bool
    errors: List[str]
    warnings: List[str] 
    stats: Dict[str, Union[int, str]]


class XLSXProcessor:
    """Excel (.xlsx) file processor for Ragnarok Online 2 CT data conversion."""
    
    # CT data type mapping for proper Excel formatting
    CT_TYPE_FORMATS: Dict[str, CTTypeFormat] = {
        'BYTE': {'num_format': '0', 'python_type': int},
        'SHORT': {'num_format': '0', 'python_type': int},
        'WORD': {'num_format': '0', 'python_type': int},
        'INT': {'num_format': '0', 'python_type': int},
        'DWORD': {'num_format': '0', 'python_type': int},
        'DWORD_HEX': {'num_format': '0', 'python_type': int},
        'INT64': {'num_format': '0', 'python_type': int},
        'FLOAT': {'num_format': '0.00', 'python_type': float},
        'BOOL': {'num_format': 'General', 'python_type': bool},
        'STRING': {'num_format': 'General', 'python_type': str}
    }

    def __init__(self, path: str, verbose: bool = False):
        """Initialize the XLSX processor."""
        if not path:
            raise ValueError("Path must be a non-empty string")
            
        self.path = Path(path)
        self.verbose = verbose
        self.header: List[str] = []
        self.types: List[str] = []
        self.rows: List[List[Any]] = []
        
        if self.verbose:
            logger.info(f"Initialized XLSX processor for: {self.path}")

    def read(self) -> List[List[str]]:
        """Parse Excel file and extract structured data."""
        try:
            if not self.path.exists():
                raise FileNotFoundError(f"Excel file not found: {self.path}")
                
            print(f"🔍 XLSX Processor: Reading file {self.path}")
            if self.verbose:
                logger.info(f"Reading Excel file: {self.path}")
            
            # Load workbook and get active sheet
            print(f"🔍 XLSX Processor: Loading workbook...")
            workbook = load_workbook(filename=str(self.path), data_only=True)
            worksheet = workbook.active
            
            if worksheet is None:
                raise ValueError("Excel file contains no active worksheet")
            
            # Read all non-empty rows
            print(f"🔍 XLSX Processor: Reading rows...")
            raw_data: List[List[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip completely empty rows
                    # Convert all cells to strings, replacing None with empty string
                    processed_row = [str(cell) if cell is not None else "" for cell in row]
                    raw_data.append(processed_row)
            
            # Validate minimum structure (header + types + at least one row)
            if len(raw_data) < 2:
                raise ValueError("Excel file must contain at least header and type rows")
              # Extract components - NEW FORMAT: types first, then headers
            # Row 1: Data types (INT, STRING, etc.)
            # Row 2: Column headers (ID, Name, etc.)  
            # Row 3+: Data rows
            self.types = raw_data[0] if len(raw_data) > 0 else []
            self.header = raw_data[1] if len(raw_data) > 1 else []
            self.rows = raw_data[2:] if len(raw_data) > 2 else []
            
            print(f"🔍 XLSX Processor: Found {len(self.types)} types, {len(self.header)} headers, {len(self.rows)} data rows")
            
            # Validate data consistency
            if len(self.types) != len(self.header):
                logger.warning(f"Type count ({len(self.types)}) doesn't match header count ({len(self.header)})")
            
            result = [self.header, self.types] + self.rows
            
            print(f"✅ XLSX Processor: Read complete - {len(self.header)} columns, {len(self.rows)} data rows")
            if self.verbose:                logger.info(f"Excel read complete: {len(self.header)} columns, {len(self.rows)} data rows")
                
            return result
            
        except InvalidFileException as e:
            print(f"💥 XLSX Processor Error: Invalid Excel file format - {e}")
            logger.error(f"Invalid Excel file format: {e}")
            raise ValueError(f"File is not a valid Excel file: {self.path}")
        except Exception as e:
            print(f"💥 XLSX Processor Error: {e}")
            logger.error(f"Error reading Excel file {self.path}: {e}")
            raise

    def write(self, data: List[List[str]]) -> None:
        """Write structured data to Excel file with proper formatting."""
        try:
            if not data or len(data) < 2:
                raise ValueError("Data must contain at least header and type rows")
            
            # Create directory if it doesn't exist
            self.path.parent.mkdir(parents=True, exist_ok=True)
            
            print(f"🔍 XLSX Processor: Writing file {self.path}")
            if self.verbose:
                logger.info(f"Writing Excel file: {self.path}")
              # Create new workbook and worksheet
            print(f"🔍 XLSX Processor: Creating workbook...")
            workbook = Workbook()
            worksheet = workbook.active
            if worksheet is None:
                raise ValueError("Failed to create worksheet")
            worksheet.title = "RO2 Table Data"
              # Extract components
            header = data[0]
            types = data[1] if len(data) > 1 else []
            rows = data[2:] if len(data) > 2 else []
            
            # Write structure: TYPE row first, then HEADER row, then DATA rows
            print(f"🔍 XLSX Processor: Writing type row...")
            self._write_type_row(worksheet, types, row_num=1)
            
            print(f"🔍 XLSX Processor: Writing header row...")
            self._write_header_row(worksheet, header, row_num=2)
            
            # Write data rows with type-aware formatting
            print(f"🔍 XLSX Processor: Writing {len(rows)} data rows...")
            self._write_data_rows(worksheet, rows, types, start_row=3)
            
            # Freeze the first two rows (types + headers)
            print(f"🔍 XLSX Processor: Freezing first two rows...")
            worksheet.freeze_panes = "A3"
            
            # Create Excel table for better data management
            print(f"🔍 XLSX Processor: Creating Excel table...")
            self._create_excel_table(worksheet, types, header, len(rows))
            
            # Apply column formatting and auto-sizing
            print(f"🔍 XLSX Processor: Formatting columns...")
            self._format_columns(worksheet, types)            # Save the workbook
            print(f"🔍 XLSX Processor: Saving workbook...")
            workbook.save(str(self.path))
            print(f"✅ XLSX Processor: Write complete - {len(header)} columns, {len(rows)} data rows")
            if self.verbose:
                logger.info(f"Excel write complete: {len(header)} columns, {len(rows)} data rows")
                
        except PermissionError as e:
            print(f"💥 XLSX Processor Error: Permission denied - {e}")
            logger.error(f"Permission denied writing to {self.path}: {e}")
            raise
        except Exception as e:
            print(f"💥 XLSX Processor Error: {e}")
            logger.error(f"Error writing Excel file {self.path}: {e}")
            raise

    def _write_header_row(self, worksheet: Worksheet, header: List[str], row_num: int = 1) -> None:
        """Write and format the header row."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, header_value in enumerate(header, 1):
            cell = worksheet.cell(row=row_num, column=col_idx, value=header_value)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _write_type_row(self, worksheet: Worksheet, types: List[str], row_num: int = 1) -> None:
        """Write and format the type definition row."""
        type_font = Font(bold=True, color="000000")
        type_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        type_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, type_value in enumerate(types, 1):
            cell = worksheet.cell(row=row_num, column=col_idx, value=type_value)
            cell.font = type_font
            cell.fill = type_fill
            cell.alignment = type_alignment

    def _write_data_rows(self, worksheet: Worksheet, rows: List[List[str]], types: List[str], start_row: int = 3) -> None:
        """Write data rows with type-appropriate value conversion."""
        for row_idx, row_data in enumerate(rows, start_row):  # Start from specified row
            for col_idx, cell_value in enumerate(row_data):
                # Convert value based on type if type information is available
                converted_value = self._convert_cell_value(cell_value, types, col_idx)
                
                cell = worksheet.cell(row=row_idx, column=col_idx + 1, value=converted_value)
                  # Apply type-specific number formatting
                if col_idx < len(types):
                    cell_type = types[col_idx]
                    if cell_type in self.CT_TYPE_FORMATS:
                        cell.number_format = self.CT_TYPE_FORMATS[cell_type]['num_format']

    def _convert_cell_value(self, value: str, types: List[str], col_idx: int) -> Any:
        """
        Convert string value to appropriate Python type based on CT type definition.
        
        Args:
            value (str): String value to convert
            types (List[str]): List of CT data types
            col_idx (int): Column index for type lookup
            
        Returns:
            Any: Converted value in appropriate Python type
        """
        if col_idx >= len(types) or not value or value.lower() == "null":
            return value
        
        cell_type = types[col_idx]
        
        try:
            if cell_type in self.CT_TYPE_FORMATS:
                target_type = self.CT_TYPE_FORMATS[cell_type]['python_type']
                
                if target_type == int:
                    # Handle hexadecimal values for DWORD_HEX type
                    if cell_type == "DWORD_HEX" and 'x' in str(value).lower():
                        return int(value, 16)  # Parse hexadecimal string
                    else:
                        return int(float(value))  # Handle cases like "1.0"
                elif target_type == float:
                    return float(value)
                elif target_type == bool:
                    return value.lower() in ('true', '1', 'yes')
                else:
                    return str(value)
            else:
                return value
                
        except (ValueError, TypeError):
            # If conversion fails, return original string value
            logger.warning(f"Could not convert '{value}' to type {cell_type}, keeping as string")
            return value

    def _format_columns(self, worksheet: Worksheet, types: List[str]) -> None:
        """Apply column formatting and auto-sizing."""
        from openpyxl.utils import get_column_letter
        
        for col_idx, _ in enumerate(types, 1):
            column_letter = get_column_letter(col_idx)
            
            # Auto-adjust column width based on content
            max_length = 0
            for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
              # Set reasonable column width (min 10, max 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def validate_structure(self, data: List[List[str]]) -> ValidationResult:
        """
        Validate the structure of data for Excel processing.
        
        Args:
            data (List[List[str]]): Data to validate
            
        Returns:
            ValidationResult: Validation results with status and details
        """
        result: ValidationResult = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'stats': {}
        }
        
        try:
            if not data:
                result['valid'] = False
                result['errors'].append("Data is empty")
                return result
            
            if len(data) < 2:
                result['valid'] = False
                result['errors'].append("Data must contain at least header and type rows")
                return result
            
            header = data[0]
            types = data[1]
            
            if len(header) != len(types):
                result['warnings'].append(f"Header count ({len(header)}) doesn't match type count ({len(types)})")
            
            # Check for empty headers
            empty_headers = [i for i, h in enumerate(header) if not h or str(h).strip() == ""]
            if empty_headers:
                result['warnings'].append(f"Empty headers found at positions: {empty_headers}")
            
            # Validate type definitions
            valid_types = set(self.CT_TYPE_FORMATS.keys())
            invalid_types = [t for t in types if t not in valid_types]
            if invalid_types:
                result['warnings'].append(f"Unknown data types: {invalid_types}")
            
            result['stats'] = {
                'columns': len(header),
                'data_rows': len(data) - 2,
                'total_rows': len(data)
            }
            
        except Exception as e:
            result['valid'] = False
            result['errors'].append(f"Validation error: {e}")
        
        return result

    def _create_excel_table(self, worksheet: Worksheet, types: List[str], header: List[str], data_rows_count: int) -> None:
        """
        Create an Excel table for better data management and presentation.
        
        Args:
            worksheet: The Excel worksheet to add the table to
            types: List of data types (will be the first row)
            header: List of column headers (will be the second row)
            data_rows_count: Number of data rows (excluding type and header rows)
        """
        try:
            # Calculate table range
            # Table includes: type row + header row + data rows
            total_rows = 2 + data_rows_count  # types + headers + data
            num_columns = len(header)
            
            if total_rows < 2 or num_columns < 1:  # Need at least 2 rows for Excel table
                print("⚠️ XLSX Processor: Cannot create table - insufficient data")
                return
            
            # Define table range starting from header row (A2) to avoid table repair issues
            # This way the type row stays separate and the table starts with actual headers
            end_column = get_column_letter(num_columns)
            table_range = f"A2:{end_column}{total_rows}"
            
            print(f"🔍 XLSX Processor: Creating table with range {table_range}")
            
            # Create table with filename as display name (safe name - no special characters)
            import re
            filename = self.path.stem  # Get filename without extension
            safe_name = re.sub(r'[^a-zA-Z0-9_]', '_', filename)
            # Ensure name starts with letter and doesn't exceed Excel limit
            if not safe_name[0].isalpha():
                safe_name = f"Table_{safe_name}"
            safe_name = safe_name[:255]  # Excel table name limit
            
            print(f"🔍 XLSX Processor: Table display name: {safe_name}")
            table = Table(displayName=safe_name, ref=table_range)
            
            # Apply table style - using a professional blue theme
            style = TableStyleInfo(
                name="TableStyleMedium2",  # Professional blue style
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            # Add the table to the worksheet
            worksheet.add_table(table)
            
            print("✅ XLSX Processor: Excel table created successfully")
            
        except Exception as e:
            print(f"⚠️ XLSX Processor: Could not create table - {e}")
            # Don't raise the exception, just log it - table creation is optional